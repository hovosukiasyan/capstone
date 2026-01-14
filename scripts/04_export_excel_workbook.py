from __future__ import annotations

from pathlib import Path
import re
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule


PROJECT_ROOT = Path(__file__).resolve().parent.parent

TIDY_DIR = PROJECT_ROOT / "data" / "processed" / "tidy"
PANEL_DIR = PROJECT_ROOT / "data" / "processed" / "panel"
OUT_DIR = PROJECT_ROOT / "data" / "processed" / "deliverables"
OUT_DIR.mkdir(parents=True, exist_ok=True)

OUT_PATH = OUT_DIR / "armstat_panel_workbook.xlsx"


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Missing file: {path}")
    return pd.read_csv(path)


def sanitize_sheet_name(name: str) -> str:
    # Excel sheet names: max 31 chars, no : \ / ? * [ ]
    name = re.sub(r"[:\\/?*\[\]]", "-", name)
    return name[:31]


def is_rate_col(col: str) -> bool:
    c = col.lower()
    return any(k in c for k in ["rate", "per_100k", "per_10k", "percap", "percent"])


def is_count_col(col: str) -> bool:
    c = col.lower()
    return any(k in c for k in ["count", "number", "total", "beds", "hospitals", "physicians", "population", "patients", "offense", "offences", "crimes"])


def apply_column_formats(ws, df: pd.DataFrame) -> None:
    """
    Apply number formats based on column names.
    """
    header = [str(c) for c in df.columns]

    # Map column index -> number format
    for idx, col in enumerate(header, start=1):
        col_l = col.lower()

        if col_l == "year":
            fmt = "0"
        elif any(x in col_l for x in ["poverty_rate", "extreme_poverty_rate", "non_poor_rate"]):
            # these are already in percent units (e.g., 27.6 meaning 27.6%)
            fmt = "0.0"
        elif "per_100k" in col_l or "per_10k" in col_l:
            fmt = "0.00"
        elif col_l == "stress_index":
            fmt = "0.000"
        elif is_count_col(col_l):
            fmt = "0"
        else:
            # default numeric
            fmt = "General"

        # Apply format to all data cells in that column
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if isinstance(cell.value, (int, float)) and fmt != "General":
                cell.number_format = fmt


def autosize_columns(ws, max_width: int = 55) -> None:
    """
    Set column widths based on max visible string length per column.
    """
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        # padding and clamp
        width = min(max_width, max(10, max_len + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_header_row(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 22

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border


def style_body(ws) -> None:
    # Light border for all cells
    thin = Side(style="thin", color="E6E6E6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alignment = Alignment(vertical="top", wrap_text=False)

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = alignment


def add_table_style(ws, table_name: str) -> None:
    """
    Wrap the sheet range into an Excel Table with banded rows + filter buttons.
    """
    if ws.max_row < 2 or ws.max_column < 1:
        return

    # Table ref (A1:...):
    last_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
    ref = f"A1:{last_cell}"

    tname = re.sub(r"[^A-Za-z0-9_]", "_", table_name)
    if len(tname) < 3:
        tname = f"Tbl_{tname}"
    tname = tname[:255]

    table = Table(displayName=tname, ref=ref)

    style = TableStyleInfo(
        name="TableStyleMedium9",  # nice blue table
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def freeze_and_filter(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_stress_conditional_formatting(ws, df: pd.DataFrame) -> None:
    """
    If sheet has stress_index column, apply a 3-color scale.
    """
    cols = [str(c) for c in df.columns]
    if "stress_index" not in cols:
        return

    idx = cols.index("stress_index") + 1
    col_letter = get_column_letter(idx)
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"

    # Green (low) -> Yellow -> Red (high)
    rule = ColorScaleRule(
        start_type="min", start_color="63BE7B",
        mid_type="percentile", mid_value=50, mid_color="FFEB84",
        end_type="max", end_color="F8696B",
    )
    ws.conditional_formatting.add(rng, rule)


def add_sheet(wb: Workbook, name: str, df: pd.DataFrame, *, table_prefix: str) -> None:
    sname = sanitize_sheet_name(name)
    ws = wb.create_sheet(title=sname)

    # Write dataframe to sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    style_header_row(ws)
    style_body(ws)
    apply_column_formats(ws, df)
    autosize_columns(ws)
    freeze_and_filter(ws)
    add_table_style(ws, table_name=f"{table_prefix}_{sname}")
    add_stress_conditional_formatting(ws, df)


def build_readme_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Sheet": [
                "README",
                "poverty_tidy",
                "crime_selected_total",
                "crime_severity_wide",
                "health_wide",
                "population_tidy",
                "panel_common_2016_2022",
                "panel_common_with_stress",
            ],
            "Description": [
                "How to use this workbook + what each sheet contains.",
                "Poverty rates by marz-year (poor/extremely poor/non-poor). Values are %.",
                "Sum of selected crime types by marz-year (from ps-ls-oc03).",
                "Crime totals + severity breakdown by marz-year (from ps-ls-oc02).",
                "Health system capacity indicators by marz-year (from ps-si-hms33).",
                "Population by marz-year (used for per-capita normalization).",
                "Merged analysis-ready panel (intersection years 2016–2022).",
                "Panel with computed Stress Index (z-score composite indicator).",
            ],
            "Notes": [
                "",
                "Keys: (marz, year).",
                "Use crime_selected_rate_per_100k for comparisons.",
                "Use crime_rate_per_100k for overall crime comparisons.",
                "Hospitals/beds are counts; per-capita columns included in panel sheets.",
                "Population is de-jure as of Jan 1.",
                "This is the main dataset for analysis/EDA/ML.",
                "Higher stress_index = higher socio-economic stress.",
            ],
        }
    )


def style_readme(ws) -> None:
    # Make README a bit nicer: larger title and wrapped text
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Wrap text in Description/Notes columns
    for row in range(2, ws.max_row + 1):
        for col in [2, 3]:
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    autosize_columns(ws, max_width=70)
    ws.freeze_panes = "A2"


def main() -> None:
    # Load datasets
    poverty = read_csv(TIDY_DIR / "poverty_tidy.csv")
    population = read_csv(TIDY_DIR / "population_tidy.csv")
    crime_selected_total = read_csv(TIDY_DIR / "crime_selected_total.csv")
    crime_severity_wide = read_csv(TIDY_DIR / "crime_severity_wide.csv")
    health_wide = read_csv(TIDY_DIR / "health_wide.csv")
    panel_common = read_csv(PANEL_DIR / "marz_year_panel_common.csv")

    stress_path = PANEL_DIR / "marz_year_panel_common_with_stress.csv"
    panel_with_stress = read_csv(stress_path) if stress_path.exists() else None

    # Create workbook
    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    # README
    readme_df = build_readme_df()
    add_sheet(wb, "README", readme_df, table_prefix="Tbl")
    style_readme(wb["README"])

    # Data sheets
    add_sheet(wb, "poverty_tidy", poverty, table_prefix="Tbl")
    add_sheet(wb, "crime_selected_total", crime_selected_total, table_prefix="Tbl")
    add_sheet(wb, "crime_severity_wide", crime_severity_wide, table_prefix="Tbl")
    add_sheet(wb, "health_wide", health_wide, table_prefix="Tbl")
    add_sheet(wb, "population_tidy", population, table_prefix="Tbl")
    add_sheet(wb, "panel_common_2016_2022", panel_common, table_prefix="Tbl")

    if panel_with_stress is not None:
        add_sheet(wb, "panel_common_with_stress", panel_with_stress, table_prefix="Tbl")

    wb.save(OUT_PATH)
    print(f"[OK] Beautified Excel workbook created: {OUT_PATH}")


if __name__ == "__main__":
    main()
